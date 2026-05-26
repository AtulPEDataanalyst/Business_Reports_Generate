"""
formatting.py — OpenPyXL formatting helpers for Insurance Processing App
Applies currency format, bold headers, auto column width, freeze pane, sheet colors.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ─── Color palette ────────────────────────────────────────────────────────────
HEADER_FILL_RAW    = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FILL_FRESH  = PatternFill("solid", fgColor="375623")   # dark green
HEADER_FILL_RENEW  = PatternFill("solid", fgColor="7B2C2C")   # dark red
TOTAL_ROW_FILL     = PatternFill("solid", fgColor="D9E1F2")   # light blue
ALT_ROW_FILL       = PatternFill("solid", fgColor="F2F2F2")   # light grey

HEADER_FONT        = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT         = Font(bold=True, size=10)
NORMAL_FONT        = Font(size=10)

CURRENCY_FORMAT    = '₹#,##0.00'
THIN_BORDER_SIDE   = Side(border_style="thin", color="BFBFBF")
THIN_BORDER        = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE
)

# Rename display headers
COLUMN_RENAME = {
    "TPOD": "Premium",
    "Total Brokerage Receivable": "Brokerage Receivable",
}

CURRENCY_COLUMNS = {
    "Premium", "Brokerage Receivable",
    "TPOD", "Total Brokerage Receivable",
    "Net Premium", "Gross Premium", "Premium Paid",
    "Basic / Own Damage Premium",
}

NUMBER_COLUMNS = {
    "Sum Insured", "New IDV",
}


def _auto_width(ws, max_width: int = 40):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _write_df_to_sheet(
    ws,
    df: pd.DataFrame,
    header_fill: PatternFill,
    freeze: bool = True,
    alternating: bool = True,
):
    """Write a DataFrame to an openpyxl worksheet with full formatting."""
    if df.empty:
        ws.append(["No data available."])
        return

    # Rename columns for display
    display_df = df.rename(columns=COLUMN_RENAME)

    # Write rows
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True)):
        ws.append(row)
        current_row = ws.max_row

        if r_idx == 0:
            # Header row
            for cell in ws[current_row]:
                cell.font = HEADER_FONT
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
        else:
            # Data row — detect if it's a "Grand Total" row
            is_total = any(
                str(cell.value).strip() in ("Grand Total", "All") 
                for cell in ws[current_row] 
                if cell.value is not None
            )

            for c_idx, cell in enumerate(ws[current_row], start=1):
                col_name = display_df.columns[c_idx - 1] if c_idx <= len(display_df.columns) else ""
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

                if is_total:
                    cell.font = TOTAL_FONT
                    cell.fill = TOTAL_ROW_FILL
                elif alternating and r_idx % 2 == 0:
                    cell.fill = ALT_ROW_FILL
                    cell.font = NORMAL_FONT
                else:
                    cell.font = NORMAL_FONT

                # Apply number formats
                if col_name in CURRENCY_COLUMNS:
                    cell.number_format = CURRENCY_FORMAT
                elif col_name in NUMBER_COLUMNS:
                    cell.number_format = '#,##0'

    # Freeze header
    if freeze:
        ws.freeze_panes = ws["A2"]

    # Row height for header
    ws.row_dimensions[1].height = 30

    _auto_width(ws)


def build_excel_output(
    raw_df: pd.DataFrame,
    fresh_df: pd.DataFrame,
    renewal_df: pd.DataFrame,
    mode: str,
) -> bytes:
    """
    Build final Excel workbook with 3 sheets:
      1. Raw Processed Data
      2. Fresh Report
      3. Renewal Report
    Returns bytes (BytesIO buffer).
    """
    fresh_title   = "Fresh Login Business"   if mode == "login" else "Fresh Issued"
    renewal_title = "Renewal Login Business" if mode == "login" else "Renewal Issued"

    wb = Workbook()

    # ── Sheet 1: Raw Processed Data ───────────────────────────────────────────
    ws_raw = wb.active
    ws_raw.title = "Raw Processed Data"
    _write_df_to_sheet(ws_raw, raw_df, HEADER_FILL_RAW, freeze=True, alternating=True)

    # ── Sheet 2: Fresh Report ──────────────────────────────────────────────────
    ws_fresh = wb.create_sheet(title=fresh_title[:31])  # sheet name max 31 chars
    _write_pivot_header(ws_fresh, fresh_title)
    # Drop internal _Report column if present
    disp_fresh = fresh_df.drop(columns=["_Report"], errors="ignore")
    _write_df_to_sheet(ws_fresh, disp_fresh, HEADER_FILL_FRESH, freeze=True, alternating=True)

    # ── Sheet 3: Renewal Report ────────────────────────────────────────────────
    ws_renew = wb.create_sheet(title=renewal_title[:31])
    _write_pivot_header(ws_renew, renewal_title)
    disp_renew = renewal_df.drop(columns=["_Report"], errors="ignore")
    _write_df_to_sheet(ws_renew, disp_renew, HEADER_FILL_RENEW, freeze=True, alternating=True)

    # Serialise to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_pivot_header(ws, title: str):
    """Write a merged title row at top of pivot sheet."""
    ws.insert_rows(1)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2E4057")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    # Merge across 10 cols (will be trimmed visually)
    try:
        ws.merge_cells("A1:J1")
    except Exception:
        pass
